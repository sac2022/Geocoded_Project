# from django.contrib.sites import requests
import json
import requests
from django.http import HttpResponse
from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
import pandas as pd
from opencage import geocoder
from opencage.geocoder import OpenCageGeocode
import openpyxl
from urllib.parse import urlencode
from django.core.files.storage import FileSystemStorage
from pandas.tests.io.excel.test_xlsxwriter import xlsxwriter
import xlsxwriter
from django.utils.encoding import smart_str
import os
from Geocoded_Project.settings import BASE_DIR


def index(request):
    # global paramas
    file_path = os.path.join(BASE_DIR, 'example.xlsx')
    api_key = "AIzaSyAxJi08HqQmW_xYO_OfmurPKQZjiwsVl4M"
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        fs = FileSystemStorage()
        name = fs.save(excel_file.name, excel_file)
        url = fs.url(name)

        print(worksheet)
        # writer=xlsxwriter.Workbook()
        # ws=writer.add_worksheet()
        # excel_sheet = wb[excel_file]
        dest_file_name = "example.xlsx"

        # book = Workbook()
        worksheet = wb.active
        excel_data = []
        # iterating over the rows and
        # getting value from each cell in row
        latlng = {}
        row_index = 1
        for row in worksheet.iter_rows():
            row_data = list()

            for cell in row:
                data_type = "json"
                parmas = {'address': str(cell.value), 'key': api_key}
                url_params = urlencode(parmas)
                endpoint = f'https://maps.googleapis.com/maps/api/geocode/{data_type}'
                url = f"{endpoint}?{url_params}"
                r = requests.get(url)
                json_data = r.json()
                try:
                    latlng = json_data['results'][0]['geometry']['location']
                except Exception as e:
                    print(e)
                lat = latlng.get('lat')
                lng = latlng.get('lng')
                c1 = worksheet.cell(row=row_index, column=2, value=lat)
                c2 = worksheet.cell(row=row_index, column=3, value=lng)
                # excel_sheet['B']  lat, lng
                print(c1.value, c2.value)
                # wb.save(excel_file)
                row_data.append(cell.value)
                row_data.append(c1.value)
                row_data.append(c2.value)
            excel_data.append(row_data)
            row_index += 1
        wb.save(filename=file_path)

        wb.close()

        # if query.status_code == "200":
        #    print(query.json())

        return render(request, 'index.html',
                      {"excel_file": excel_file, "excel_data": excel_data})  # {"excel_data": excel_data})


def download(request):
    file_path = os.path.join(BASE_DIR, 'example.xlsx')
    print(file_path)
    if os.path.exists(file_path):
        with open(file_path, "rb") as excel:
            data = excel.read()

    response = HttpResponse(data,
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % smart_str(file_path)
    response['X-Sendfile'] = smart_str(file_path)
    return response
